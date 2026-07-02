#!/usr/bin/env python3
"""Export ranked submission to XLSX for portal upload."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from src.utils import DATA_RAW, OUTPUTS_DIR


def export_xlsx(
    csv_path: Path,
    candidates_path: Path,
    out_path: Path,
) -> None:
    rows = list(csv.DictReader(open(csv_path, encoding="utf-8")))
    ids_needed = {r["candidate_id"] for r in rows}
    names: dict[str, str] = {}

    with open(candidates_path, encoding="utf-8") as f:
        for line in f:
            d = json.loads(line)
            cid = d["candidate_id"]
            if cid in ids_needed:
                names[cid] = d["profile"].get("anonymized_name", "")
                if len(names) == len(ids_needed):
                    break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 100 Candidates"

    headers = [
        "Rank",
        "Candidate_ID",
        "Candidate_Name",
        "Final_Score",
        "Confidence",
        "Recommendation_Reason",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        score = float(r["score"])
        ws.append([
            int(r["rank"]),
            r["candidate_id"],
            names.get(r["candidate_id"], ""),
            score,
            round(max(0.35, min(0.99, score)), 3),
            r["reasoning"],
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    ws.column_dimensions["F"].width = 80
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export submission CSV to XLSX")
    parser.add_argument("--csv", default=str(OUTPUTS_DIR / "submission.csv"))
    parser.add_argument("--candidates", default=str(DATA_RAW / "candidates.jsonl"))
    parser.add_argument("--out", default=str(OUTPUTS_DIR / "submission.xlsx"))
    args = parser.parse_args()
    export_xlsx(Path(args.csv), Path(args.candidates), Path(args.out))
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
